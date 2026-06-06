#!/usr/bin/env python3
"""
Combo → Supabase importer
=========================

Importe les pointages depuis les fichiers Excel exportés de Combo
(format "Rapport des émargements et validations - YYYY-MM-DD au YYYY-MM-DD.xlsx")
vers la table `combo_pointages` de Supabase.

Usage :
    python3 combo_import.py <fichier.xlsx> [<fichier2.xlsx> ...]
    python3 combo_import.py --all <dossier>      # Tous les .xlsx d'un dossier
    python3 combo_import.py --dry-run <fichier>  # Simule sans insérer

Variables d'environnement requises :
    SUPABASE_URL          ex: https://dzrherfavgiuygnimtux.supabase.co
    SUPABASE_SERVICE_KEY  service_role key (PAS l'anon key, sinon RLS bloque)

L'import est idempotent : la contrainte UNIQUE(date, collaborateur, etab, debut_planifie)
permet de réimporter le même fichier sans créer de doublons (UPSERT via ON CONFLICT).

Auteur : session Cowork PlanB-Tools, 2026-06-06
"""

import os
import sys
import argparse
import logging
from datetime import datetime, date, time, timedelta, timezone
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("ERREUR : module openpyxl manquant. Installer avec : pip3 install openpyxl")
    sys.exit(1)

try:
    from supabase import create_client
except ImportError:
    print("ERREUR : module supabase manquant. Installer avec : pip3 install supabase")
    sys.exit(1)


# ============================================================
# CONFIGURATION
# ============================================================

PARIS_TZ_OFFSET_HOURS_SUMMER = 2  # CEST (été)
PARIS_TZ_OFFSET_HOURS_WINTER = 1  # CET (hiver)

# Mapping établissements Combo → PlanB-Tools (en dur car peu de chances de changer)
ETAB_MAPPING = {
    'chez  l oncle freddy': 'freddy',   # 2 espaces, sans apostrophe (exact Combo)
    'chez l oncle freddy': 'freddy',     # variante 1 espace au cas où
    'chez tante liesel': 'liesel',
}

# Mapping équipes Combo → PlanB-Tools (normalisation)
EQUIPE_MAPPING = {
    'salle': 'salle',
    'cuisine': 'cuisine',
    'aide cuisine / plonge': 'plonge',
    'plonge': 'plonge',
}

# Noms des colonnes attendues (par index) dans le fichier Combo
COL_JOUR = 0
COL_ETAB = 1
COL_EQUIPE = 2
COL_COLLAB = 3
COL_DEBUT_PLAN = 4
COL_FIN_PLAN = 5
COL_PAUSE_PLAN = 6
COL_DEBUT_POINTE = 7
COL_FIN_POINTEE = 8
COL_PAUSE_POINTEE = 9
COL_DEBUT_VALIDE = 10
COL_FIN_VALIDEE = 11
COL_PAUSE_VALIDEE = 12
COL_VALIDE_PAR = 13
COL_COMM_DEBUT = 14
COL_COMM_FIN = 15
COL_COMM_VAL = 16


# ============================================================
# HELPERS
# ============================================================

def is_dst_paris(d: date) -> bool:
    """Approximation simple : DST en France de fin mars à fin octobre."""
    return 3 < d.month < 11 or (d.month == 3 and d.day >= 27) or (d.month == 10 and d.day < 27)


def to_paris_tz(d: date, t: time) -> datetime:
    """Combine date + heure en UTC en partant du principe que l'heure est en heure de Paris."""
    if t is None:
        return None
    offset = PARIS_TZ_OFFSET_HOURS_SUMMER if is_dst_paris(d) else PARIS_TZ_OFFSET_HOURS_WINTER
    naive = datetime.combine(d, t)
    # naive est en heure de Paris → on soustrait l'offset pour avoir UTC
    return (naive - timedelta(hours=offset)).replace(tzinfo=timezone.utc)


def build_timestamps(d_service: date, t_debut: time, t_fin: time) -> tuple:
    """
    Combine date + heures début/fin en timestamps UTC.
    Gère le passage minuit : si fin < début, fin est sur date_service + 1 jour.
    """
    if t_debut is None and t_fin is None:
        return None, None
    debut_ts = to_paris_tz(d_service, t_debut) if t_debut else None
    if t_fin is None:
        return debut_ts, None
    if t_debut is not None and t_fin <= t_debut:
        # Passage minuit
        fin_ts = to_paris_tz(d_service + timedelta(days=1), t_fin)
    else:
        fin_ts = to_paris_tz(d_service, t_fin)
    return debut_ts, fin_ts


def parse_date(raw) -> date:
    """Combo donne la date soit en datetime, soit en string 'DD/MM/YYYY'."""
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, str):
        for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
            try:
                return datetime.strptime(raw.strip(), fmt).date()
            except ValueError:
                continue
    raise ValueError(f"Format de date inattendu: {raw!r}")


def parse_time(raw) -> time:
    """Combo donne l'heure soit en time, soit en string 'HH:MM'. None possible."""
    if raw is None or raw == '':
        return None
    if isinstance(raw, time):
        return raw
    if isinstance(raw, datetime):
        return raw.time()
    if isinstance(raw, str):
        for fmt in ('%H:%M', '%H:%M:%S'):
            try:
                return datetime.strptime(raw.strip(), fmt).time()
            except ValueError:
                continue
    raise ValueError(f"Format d'heure inattendu: {raw!r}")


def normalize_etab(etab_combo: str) -> str:
    """Mappe le nom Combo vers freddy/liesel."""
    key = etab_combo.lower().strip()
    return ETAB_MAPPING.get(key, key)


def normalize_equipe(equipe_combo: str) -> str:
    """Normalise le nom de l'équipe."""
    if not equipe_combo:
        return None
    key = equipe_combo.lower().strip()
    return EQUIPE_MAPPING.get(key, key)


# ============================================================
# IMPORT
# ============================================================

def load_user_mapping(client) -> dict:
    """Charge la table combo_user_mapping en cache local."""
    res = client.table('combo_user_mapping').select('combo_collaborateur_nom, user_id').execute()
    return {row['combo_collaborateur_nom']: row['user_id'] for row in res.data}


def parse_excel_file(xlsx_path: Path) -> list:
    """Parse un fichier Excel Combo et retourne une liste de dicts (1 par ligne)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if 'Planb' not in wb.sheetnames:
        raise ValueError(f"Onglet 'Planb' introuvable dans {xlsx_path.name}")
    ws = wb['Planb']

    rows = []
    for row_idx, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not raw[COL_COLLAB]:
            continue  # ligne vide
        try:
            d_service = parse_date(raw[COL_JOUR])
            t_dp = parse_time(raw[COL_DEBUT_PLAN])
            t_fp = parse_time(raw[COL_FIN_PLAN])
            t_dpt = parse_time(raw[COL_DEBUT_POINTE])
            t_fpt = parse_time(raw[COL_FIN_POINTEE])
            t_dv = parse_time(raw[COL_DEBUT_VALIDE])
            t_fv = parse_time(raw[COL_FIN_VALIDEE])

            debut_plan, fin_plan = build_timestamps(d_service, t_dp, t_fp)
            debut_pt, fin_pt = build_timestamps(d_service, t_dpt, t_fpt)
            debut_val, fin_val = build_timestamps(d_service, t_dv, t_fv)

            etab_brut = raw[COL_ETAB] or ''
            equipe_brut = raw[COL_EQUIPE] or ''

            rows.append({
                'date_service': d_service.isoformat(),
                'etablissement_combo': etab_brut,
                'etablissement': normalize_etab(etab_brut),
                'equipe_combo': equipe_brut,
                'equipe': normalize_equipe(equipe_brut),
                'collaborateur_nom_combo': raw[COL_COLLAB],
                # user_id sera rempli par lookup ensuite
                'debut_planifie': debut_plan.isoformat() if debut_plan else None,
                'fin_planifiee': fin_plan.isoformat() if fin_plan else None,
                'pauses_planifiees_minutes': int(raw[COL_PAUSE_PLAN] or 0),
                'debut_pointe': debut_pt.isoformat() if debut_pt else None,
                'fin_pointee': fin_pt.isoformat() if fin_pt else None,
                'pauses_pointees_minutes': int(raw[COL_PAUSE_POINTEE] or 0),
                'debut_valide': debut_val.isoformat() if debut_val else None,
                'fin_validee': fin_val.isoformat() if fin_val else None,
                'pauses_validees_minutes': int(raw[COL_PAUSE_VALIDEE] or 0),
                'valide_par_nom': raw[COL_VALIDE_PAR],
                'commentaire_debut_pointe': raw[COL_COMM_DEBUT],
                'commentaire_fin_pointee': raw[COL_COMM_FIN],
                'commentaire_validation': raw[COL_COMM_VAL],
                'import_fichier_nom': xlsx_path.name,
            })
        except Exception as e:
            logging.error(f"Erreur ligne {row_idx} de {xlsx_path.name}: {e} — ligne ignorée")
            continue
    return rows


def import_file(client, xlsx_path: Path, user_mapping: dict, dry_run: bool = False) -> dict:
    """Importe un fichier Combo. Retourne un rapport."""
    logging.info(f"📄 Lecture {xlsx_path.name}")
    rows = parse_excel_file(xlsx_path)
    logging.info(f"   {len(rows)} lignes parsées")

    # Lookup user_id
    rows_with_user = 0
    rows_without_user = []
    for r in rows:
        uid = user_mapping.get(r['collaborateur_nom_combo'])
        r['user_id'] = uid
        if uid:
            rows_with_user += 1
        else:
            rows_without_user.append(r['collaborateur_nom_combo'])

    logging.info(f"   {rows_with_user} lignes avec user_id mappé")
    if rows_without_user:
        unique = set(rows_without_user)
        logging.warning(f"   ⚠️  {len(rows_without_user)} lignes sans user_id mappé "
                        f"({len(unique)} noms uniques): {sorted(unique)}")

    if dry_run:
        logging.info("   [DRY-RUN] aucune insertion en base")
        return {
            'fichier': xlsx_path.name,
            'lignes_parsees': len(rows),
            'lignes_avec_user': rows_with_user,
            'lignes_sans_user': len(rows_without_user),
            'inserees_ou_mises_a_jour': 0,
            'dry_run': True,
        }

    # UPSERT par batch
    upserted = 0
    BATCH = 100
    for i in range(0, len(rows), BATCH):
        batch = rows[i:i+BATCH]
        try:
            res = client.table('combo_pointages').upsert(
                batch,
                on_conflict='date_service,collaborateur_nom_combo,etablissement_combo,debut_planifie'
            ).execute()
            upserted += len(res.data) if res.data else len(batch)
        except Exception as e:
            logging.error(f"   ❌ Erreur batch {i}-{i+len(batch)}: {e}")

    logging.info(f"   ✅ {upserted} lignes upsertées (insérées ou mises à jour)")

    return {
        'fichier': xlsx_path.name,
        'lignes_parsees': len(rows),
        'lignes_avec_user': rows_with_user,
        'lignes_sans_user': len(rows_without_user),
        'inserees_ou_mises_a_jour': upserted,
        'noms_non_mappes': sorted(set(rows_without_user)),
    }


# ============================================================
# CLI
# ============================================================

def main():
    parser = argparse.ArgumentParser(description='Import Combo → Supabase')
    parser.add_argument('files', nargs='*', help='Fichiers .xlsx à importer')
    parser.add_argument('--all', metavar='DIR', help='Importer tous les .xlsx d\'un dossier')
    parser.add_argument('--dry-run', action='store_true', help='Simuler sans insérer')
    parser.add_argument('-v', '--verbose', action='store_true', help='Logs détaillés')
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format='%(asctime)s %(levelname)s %(message)s',
        datefmt='%H:%M:%S'
    )

    # Configuration Supabase
    url = os.environ.get('SUPABASE_URL')
    key = os.environ.get('SUPABASE_SERVICE_KEY')
    if not url or not key:
        logging.error("SUPABASE_URL et SUPABASE_SERVICE_KEY doivent être définis")
        sys.exit(1)
    client = create_client(url, key)

    # Liste des fichiers
    files = []
    if args.all:
        files = sorted(Path(args.all).glob('*.xlsx'))
    if args.files:
        files.extend(Path(f) for f in args.files)
    if not files:
        logging.error("Aucun fichier à importer. Donner des fichiers en arg ou --all <dossier>")
        sys.exit(1)

    # Charge le mapping users
    logging.info("📋 Chargement du mapping users...")
    user_mapping = load_user_mapping(client)
    logging.info(f"   {len(user_mapping)} mappings chargés ({sum(1 for v in user_mapping.values() if v)} avec user_id)")

    # Import
    reports = []
    for f in files:
        if not f.exists():
            logging.warning(f"⚠️  Fichier introuvable: {f}")
            continue
        report = import_file(client, f, user_mapping, dry_run=args.dry_run)
        reports.append(report)

    # Récap global
    print("\n" + "="*70)
    print("📊 RÉCAP GLOBAL")
    print("="*70)
    total_parsees = sum(r['lignes_parsees'] for r in reports)
    total_inserees = sum(r['inserees_ou_mises_a_jour'] for r in reports)
    total_sans_user = sum(r['lignes_sans_user'] for r in reports)
    all_noms_non_mappes = set()
    for r in reports:
        all_noms_non_mappes.update(r.get('noms_non_mappes', []))

    print(f"  Fichiers traités       : {len(reports)}")
    print(f"  Lignes parsées         : {total_parsees}")
    print(f"  Lignes upsertées en DB : {total_inserees}")
    print(f"  Lignes sans user_id    : {total_sans_user} ({100*total_sans_user//max(total_parsees,1)}%)")
    if all_noms_non_mappes:
        print(f"\n  Noms Combo non mappés (à ajouter dans combo_user_mapping si nécessaire) :")
        for n in sorted(all_noms_non_mappes):
            print(f"    - {n}")
    print()


if __name__ == '__main__':
    main()
